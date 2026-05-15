# -*- coding: utf-8 -*-
"""
validar_excel_generado.py — MejorAhora SAS (2026-04-23)
========================================================
Validador M2 post-generacion Excel. Lee el .xlsx producido por el
populator y verifica celdas canonicas contra valores esperados.

Objetivo: detectar regresiones sin abrir Excel manualmente. Si falla,
el pipeline agrega warning al log; el analista revisa visualmente.

Validaciones en hoja ACTUAL:
  - B2  (credito)           == datos.credito_id
  - B5  (cuota)             ≈ datos.cuota_mensual (tol ±$1)
  - B7  (plazo pendiente)   == datos.plazo_pendiente (en meses)
  - B10 (seguro_vida)       ≈ datos.seguro_vida
  - B11 (seguro_incendio)   ≈ datos.seguro_incendio
  - B12 (seguro_terremoto)  ≈ datos.seguro_terremoto
  - B15 (saldo capital)     ≈ datos.saldo_capital
  - 6 opciones (B16:B21)    plazos en orden DESCENDENTE (R-3b)
  - activeTab               == 0 (hoja ACTUAL seleccionada)

NOTA: B13 (capital_mensual) y B14 (interes_mensual) NO se validan en M2
porque la Regla 9.3 los ajusta post-populator y el valor canonico es el
del Excel ya escrito (la cuota se reconstituye por columnas). El control
de coherencia 9.3 vive en M1 (validar_extraccion_davivienda) antes del
write, y en la suma vs cuota que la propia formula B5 reproduce.

Filename: ESTUDIO <NOMBRE MAYUSCULAS>-<CCCC>-DD.MM.AA.xlsx (2026-05-15 Jose feedback).
CCCC = ultimos 4 digitos del credito (antes del guion verificador). Permite
distinguir multiples creditos del mismo cliente.

Uso programatico:
    from validar_excel_generado import validar_excel
    ok, errores, warnings = validar_excel(path_xlsx, datos)
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


TOLERANCIA_PESOS = 1.0  # Excel redondea, tol estricta de $1


def _aprox(a, b, tol=TOLERANCIA_PESOS) -> bool:
    try:
        return abs(float(a or 0) - float(b or 0)) <= tol
    except (ValueError, TypeError):
        return False


def _check(errores, cond, msg):
    if not cond:
        errores.append(msg)


def validar_excel(path_xlsx, datos) -> tuple[bool, list[str], list[str]]:
    """Valida Excel generado contra DatosClienteExcel. Retorna (ok, errores, warnings)."""
    errores: list[str] = []
    warnings: list[str] = []

    path = Path(path_xlsx)
    if not path.exists():
        errores.append(f"Excel no existe: {path}")
        return False, errores, warnings

    # 1) Naming: ESTUDIO <NOMBRE>-<CCCC>-DD.MM.AA.xlsx
    # 2026-05-15 (Jose feedback): nuevo formato con credito_corto (4 digitos)
    # para distinguir multiples creditos del mismo cliente.
    nombre_esperado_fragment = (datos.nombre or "").strip()
    name = path.name
    if not name.startswith("ESTUDIO "):
        errores.append(f"naming invalido (no empieza con 'ESTUDIO '): {name}")
    if nombre_esperado_fragment and nombre_esperado_fragment not in name.upper():
        warnings.append(
            f"nombre cliente '{nombre_esperado_fragment}' no aparece en filename: {name}"
        )
    # Patron nuevo: -CCCC-DD.MM.AA.xlsx
    # Patron viejo (retrocompat): -DD.MM.AA.xlsx
    patron_nuevo = re.search(r"-(\d{4})-(\d{2}\.\d{2}\.\d{2})\.xlsx$", name)
    patron_viejo = re.search(r"-(\d{2}\.\d{2}\.\d{2})\.xlsx$", name)
    if not patron_nuevo and not patron_viejo:
        warnings.append(
            f"filename sin sufijo fecha valido (esperado -CCCC-DD.MM.AA.xlsx): {name}"
        )
    elif patron_viejo and not patron_nuevo:
        # Excel del flujo viejo (sin credito_corto). Aceptable retrocompat pero avisar.
        warnings.append(
            f"filename formato viejo (sin credito_corto). Se acepta retrocompat: {name}"
        )

    try:
        wb = load_workbook(str(path), data_only=False, keep_vba=False)
    except Exception as e:
        errores.append(f"no se pudo abrir xlsx: {e}")
        return False, errores, warnings

    # 2) Hoja ACTUAL debe existir
    if "ACTUAL" not in wb.sheetnames:
        errores.append(f"hoja 'ACTUAL' no existe. Hojas: {wb.sheetnames}")
        return False, errores, warnings

    ws = wb["ACTUAL"]

    # 3) activeTab == 0 (ACTUAL seleccionada) — feedback_excel_copiar_celdas
    # (no leemos sheet_view.tabSelected porque wb.index(wb.active) ya cubre el caso)
    if wb.index(wb.active) != 0:
        warnings.append(
            f"activeTab != 0 (ACTUAL deberia ser la primera seleccionada). "
            f"Actual: {wb.active.title}"
        )

    # 4) Valores en celdas canonicas (hoja ACTUAL)
    _check(errores,
           str(ws["B2"].value or "").strip() == (datos.credito_id or "").strip(),
           f"B2 credito_id: esperado={datos.credito_id!r} got={ws['B2'].value!r}")
    _check(errores, _aprox(ws["B5"].value, datos.cuota_mensual),
           f"B5 cuota: esperado=${datos.cuota_mensual:,.0f} got={ws['B5'].value!r}")
    # B7 plazo_pendiente (entero, meses) — promesa del docstring que faltaba implementar
    try:
        b7_int = int(float(ws["B7"].value)) if ws["B7"].value is not None else None
    except (ValueError, TypeError):
        b7_int = None
    _check(errores,
           b7_int is not None and b7_int == int(datos.plazo_pendiente or 0),
           f"B7 plazo_pendiente: esperado={int(datos.plazo_pendiente or 0)} "
           f"got={ws['B7'].value!r}")
    _check(errores, _aprox(ws["B10"].value, datos.seguro_vida),
           f"B10 seguro_vida: esperado=${datos.seguro_vida:,.0f} got={ws['B10'].value!r}")
    _check(errores, _aprox(ws["B11"].value, datos.seguro_incendio),
           f"B11 seguro_incendio: esperado=${datos.seguro_incendio:,.0f} got={ws['B11'].value!r}")
    _check(errores, _aprox(ws["B12"].value, datos.seguro_terremoto),
           f"B12 seguro_terremoto: esperado=${datos.seguro_terremoto:,.0f} got={ws['B12'].value!r}")
    _check(errores, _aprox(ws["B15"].value, datos.saldo_capital),
           f"B15 saldo_capital: esperado=${datos.saldo_capital:,.0f} got={ws['B15'].value!r}")

    # 5) Plazos en B16:B21 en orden DESCENDENTE (R-3b)
    plazos_excel = []
    for row in range(16, 22):
        v = ws[f"B{row}"].value
        if isinstance(v, (int, float)):
            plazos_excel.append(float(v))
    if len(plazos_excel) < 6:
        warnings.append(
            f"menos de 6 opciones en B16:B21 (got {len(plazos_excel)}): {plazos_excel}"
        )
    elif len(plazos_excel) == 6:
        for i in range(5):
            if plazos_excel[i] < plazos_excel[i+1]:
                errores.append(
                    f"orden opciones QUEBRADO (R-3b): B{16+i}={plazos_excel[i]} "
                    f"< B{17+i}={plazos_excel[i+1]}. Todas deben ser descendentes."
                )
                break

    ok = len(errores) == 0
    return ok, errores, warnings


def formatear_reporte(ok: bool, errores: list[str], warnings: list[str]) -> str:
    lines = []
    lines.append(f"[M2-validar-xlsx] {'OK' if ok else 'FAIL'} "
                 f"({len(errores)} errores, {len(warnings)} warnings)")
    for e in errores:
        lines.append(f"  ERROR: {e}")
    for w in warnings:
        lines.append(f"  WARN:  {w}")
    return "\n".join(lines)
