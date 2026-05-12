"""
Auto-Poblador de Excel - MejorAhora SAS (v2.0 - ZIP directo)
==============================================================
Manipula el XLSX a nivel ZIP/XML para preservar formato exacto.
NO usa openpyxl para guardar (destruye drawings/bordes).

Version: 2.0
"""

from __future__ import annotations
import os, shutil, re, zipfile, tempfile, subprocess
from datetime import datetime
from dataclasses import dataclass
from xml.etree import ElementTree as ET
from copy import deepcopy

# Namespace del spreadsheetML
NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

# URI -> prefijo canonico que Excel espera. Clave '' = namespace default (sin prefijo).
_URI_TO_PREFIX = {
    NS: '',  # spreadsheetml main -> default (sin prefijo)
    NS_R: 'r',
    'http://schemas.openxmlformats.org/markup-compatibility/2006': 'mc',
    'http://schemas.microsoft.com/office/spreadsheetml/2009/9/main': 'x14',
    'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac': 'x14ac',
    'http://schemas.microsoft.com/office/spreadsheetml/2010/11/main': 'x15',
    'http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac': 'x15ac',
    'http://schemas.microsoft.com/office/spreadsheetml/2014/revision': 'xr',
    'http://schemas.microsoft.com/office/spreadsheetml/2015/revision2': 'xr2',
    'http://schemas.microsoft.com/office/spreadsheetml/2016/revision6': 'xr6',
    'http://schemas.microsoft.com/office/spreadsheetml/2016/revision10': 'xr10',
    'http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures': 'xcalcf',
    'http://schemas.microsoft.com/office/excel/2006/main': 'xne',
    'http://schemas.openxmlformats.org/officeDocument/2006/sharedTypes': 'st',
}
# Registrar para que ET los use desde el inicio cuando sea posible
for _uri, _prefix in _URI_TO_PREFIX.items():
    ET.register_namespace(_prefix, _uri)


def _fix_xml_namespaces(xml_bytes):
    """Post-procesa XML generado por ET.tostring para normalizar prefijos ns0:, ns1:...

    ET.tostring a veces genera ns0:, ns4: etc. cuando el XML original tiene URIs
    que no fueron pre-registrados. Esto rompe Excel en modo estricto.
    Esta funcion:
      1) Encuentra todos los xmlns:nsN="URI" en el root
      2) Reemplaza nsN: por el prefijo canonico de ese URI (o lo elimina si es el default)
      3) Corrige las declaraciones xmlns para usar prefijos canonicos
      4) Normaliza la declaracion XML al formato que Excel espera:
         <?xml version="1.0" encoding="UTF-8" standalone="yes"?>
         (comillas dobles + standalone="yes")
    """
    xml = xml_bytes.decode('utf-8') if isinstance(xml_bytes, bytes) else xml_bytes

    # Normalizar declaracion XML: forzar comillas dobles + standalone="yes"
    # (Excel estricto rechaza comillas simples y requiere standalone)
    xml = re.sub(
        r"<\?xml[^?]*\?>",
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        xml,
        count=1,
    )
    # Si no habia declaracion, agregarla
    if not xml.lstrip().startswith('<?xml'):
        xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml.lstrip()

    # Extraer declaraciones de namespace del root (xmlns:nsN="URI")
    # Busca en el primer elemento root abierto
    root_match = re.search(r'<([a-zA-Z0-9]+:)?([a-zA-Z0-9]+)([^>]*)>', xml)
    if not root_match:
        return xml_bytes if isinstance(xml_bytes, bytes) else xml.encode('utf-8')

    root_attrs = root_match.group(3)
    # Encontrar todos los xmlns:prefijo="URI"
    ns_decls = re.findall(r'xmlns:([a-zA-Z0-9_]+)="([^"]+)"', root_attrs)
    # Tambien el default xmlns="URI"
    default_ns_match = re.search(r'\bxmlns="([^"]+)"', root_attrs)

    # Mapa: prefijo_actual -> prefijo_canonico (o '' si default)
    rename_map = {}
    for curr_prefix, uri in ns_decls:
        if uri in _URI_TO_PREFIX:
            canonical = _URI_TO_PREFIX[uri]
            if canonical != curr_prefix:
                rename_map[curr_prefix] = canonical

    # Aplicar los renames: cada `nsN:` se reemplaza por el canonico
    for curr_prefix, canonical in rename_map.items():
        if canonical == '':
            # Convertir a default: elimina "nsN:" de elementos y atributos
            xml = re.sub(r'\b' + re.escape(curr_prefix) + r':', '', xml)
        else:
            xml = re.sub(r'\b' + re.escape(curr_prefix) + r':', canonical + ':', xml)

    # Corregir declaraciones xmlns en el root:
    #   - xmlns:nsN="URI"   -> xmlns:canonical="URI"   (si canonical != '')
    #   - xmlns:nsN="URI"   -> xmlns="URI"             (si canonical == '')
    def _fix_xmlns_decl(m):
        full_attrs = m.group(3)
        pre, tag = m.group(1) or '', m.group(2)

        def _repl_decl(d):
            prefix = d.group(1)
            uri = d.group(2)
            if uri in _URI_TO_PREFIX:
                canonical = _URI_TO_PREFIX[uri]
                if canonical == '':
                    return f'xmlns="{uri}"'
                else:
                    return f'xmlns:{canonical}="{uri}"'
            return d.group(0)

        new_attrs = re.sub(r'xmlns:([a-zA-Z0-9_]+)="([^"]+)"', _repl_decl, full_attrs)
        # Eliminar duplicados de xmlns="URI" si quedaron
        seen = set()
        def _dedup(d):
            decl = d.group(0)
            if decl in seen:
                return ''
            seen.add(decl)
            return decl
        new_attrs = re.sub(r'xmlns(?::[a-zA-Z0-9_]+)?="[^"]+"', _dedup, new_attrs)
        return f'<{pre}{tag}{new_attrs}>'

    xml = re.sub(r'<([a-zA-Z0-9]+:)?([a-zA-Z0-9]+)([^>]*)>', _fix_xmlns_decl, xml, count=1)

    # PASO FINAL: asegurar que los prefijos usados en mc:Ignorable esten
    # declarados en el root. Si no lo estan, Excel estricto rechaza el XML.
    # Mapa conocido de prefijos comunes a sus URIs:
    _PREFIX_TO_URI = {
        'x14': 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/main',
        'x14ac': 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac',
        'x15': 'http://schemas.microsoft.com/office/spreadsheetml/2010/11/main',
        'x15ac': 'http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac',
        'xr': 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision',
        'xr2': 'http://schemas.microsoft.com/office/spreadsheetml/2015/revision2',
        'xr3': 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision3',
        'xr6': 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision6',
        'xr10': 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision10',
        'xcalcf': 'http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures',
        'xne': 'http://schemas.microsoft.com/office/excel/2006/main',
    }

    # Escanear TODO el XML para encontrar prefijos usados.
    # Solo cuentan los que aparecen como nombre de elemento (<prefijo:tag) o
    # atributo (espacio prefijo:attr=). Asi evitamos capturar referencias de
    # celdas de Excel como "B5:B10" o rutas como "C:\..." o URIs "http://...".
    used_prefixes = set()
    # 1) Nombres de elemento: <prefijo:tag o </prefijo:tag
    used_prefixes.update(re.findall(r'</?([a-zA-Z][a-zA-Z0-9_]*):[a-zA-Z]', xml))
    # 2) Atributos con prefijo: (espacio) prefijo:attr=
    used_prefixes.update(re.findall(r'\s([a-zA-Z][a-zA-Z0-9_]*):[a-zA-Z][a-zA-Z0-9_]*\s*=', xml))
    # 3) Tambien los listados en mc:Ignorable
    for ig in re.finditer(r'mc:Ignorable="([^"]+)"', xml):
        for p in ig.group(1).split():
            used_prefixes.add(p)
    # Excluir keywords XML
    used_prefixes.discard('xml')
    used_prefixes.discard('xmlns')

    def _ensure_all_prefixes(m):
        attrs = m.group(3)
        existing = set(re.findall(r'xmlns:([a-zA-Z0-9_]+)=', attrs))
        # Tambien el default namespace cuenta
        to_add = []
        for p in used_prefixes:
            if p not in existing and p in _PREFIX_TO_URI:
                to_add.append('xmlns:' + p + '="' + _PREFIX_TO_URI[p] + '"')
        if to_add:
            # Insertar antes de mc:Ignorable si existe, sino al inicio de attrs
            ig_pos = attrs.find('mc:Ignorable')
            insert_str = ' '.join(to_add) + ' '
            if ig_pos > 0:
                new_attrs = attrs[:ig_pos] + insert_str + attrs[ig_pos:]
            else:
                new_attrs = attrs.rstrip() + ' ' + insert_str.rstrip() + (attrs[len(attrs.rstrip()):] if attrs != attrs.rstrip() else '')
            pre, tag = m.group(1) or '', m.group(2)
            return '<' + pre + tag + new_attrs + '>'
        return m.group(0)

    xml = re.sub(r'<([a-zA-Z0-9]+:)?(worksheet|workbook|Relationships|Types)([^>]*)>',
                 _ensure_all_prefixes, xml, count=1)

    return xml.encode('utf-8')


def to_float(val):
    """Convierte valor a float limpiando formato ($, comas, %)."""
    if val is None or val == '' or val == 0:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('$', '').replace(',', '').replace(' ', '').strip()
    if s.endswith('%'):
        try:
            return float(s[:-1]) / 100
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def to_int(val):
    return int(to_float(val))


# Celdas INDEX/MATCH en ACTUAL (sheet2.xml) que se reemplazan
INDEX_CELLS = {
    'B3': ('str', 'nombre'),
    'B4': ('str', 'banco'),
    'B5': ('num', 'cuota_mensual'),
    'B6': ('num', 'plazo_inicial'),
    'B7': ('num', 'plazo_pendiente'),
    'B9': ('num', 'frech_subsidio'),
    'B10': ('num', 'seguro_vida'),
    'B11': ('num', 'seguro_incendio'),
    'B12': ('num', 'seguro_terremoto'),
    'B13': ('num', 'capital_mensual'),
    'B14': ('num', 'interes_mensual'),
    'B15': ('num', 'saldo_capital'),
    'B22': ('str', 'consultor'),
    'B23': ('str', 'actividad_economica'),
    'B25': ('num', 'abono_efectivo'),
    'B26': ('num', 'ingresos'),
}

# Celdas de valor directo (sin formula INDEX)
DIRECT_CELLS = {
    'B2': ('str', 'credito_id'),
    'B8': ('num', 'tasa_ea'),
}

PLAZO_CELLS = ['B16', 'B17', 'B18', 'B19', 'B20', 'B21']

# Plazos por defecto (decrecientes, en anos). Importado de reglas_negocio
# para evitar duplicacion (2026-05-12: antes era literal local stale-prone).
from reglas_negocio import PLAZOS_DEFAULT as _PLAZOS_DEFAULT  # noqa: E402


def _calcular_plazos_dinamicos(plazo_pendiente_meses):
    """Devuelve 6 plazos en anos, TODOS estrictamente menores al plazo pendiente.

    - Si plazo_pendiente >= 14 anos: usa los defaults [13.5, 12, 11, 10, 9, 8.5]
    - Si plazo_pendiente < 14 anos: genera escalonado decreciente desde
      el mayor entero menor a plazo_pendiente, en pasos de 1 ano,
      con un minimo de 4 anos.

    Ejemplos:
      plazo_pendiente 240m (20 anos) -> [13.5, 12, 11, 10, 9, 8.5]
      plazo_pendiente 118m (9.83 anos) -> [9, 8, 7, 6, 5, 4]
      plazo_pendiente 100m (8.33 anos) -> [8, 7, 6, 5, 4, 4]  (clamp en 4)
      plazo_pendiente 60m (5 anos) -> [4, 4, 4, 4, 4, 4]
    """
    if plazo_pendiente_meses <= 0:
        return _PLAZOS_DEFAULT.copy()

    plazo_pend_anos = plazo_pendiente_meses / 12.0

    # Caso 1: defaults caben todos
    if plazo_pend_anos > 13.5:
        return _PLAZOS_DEFAULT.copy()

    # Caso 2: generar escalonado desde el entero menor al pendiente
    import math
    techo = math.floor(plazo_pend_anos)  # mayor entero < plazo_pend si plazo_pend no es entero
    if techo == plazo_pend_anos:
        techo -= 1  # estrictamente menor
    techo = max(techo, 4)  # minimo 4 anos

    plazos = []
    actual = techo
    for _ in range(6):
        plazos.append(float(max(actual, 4)))
        actual -= 1
    return plazos


@dataclass
class DatosClienteExcel:
    credito_id: str
    nombre: str
    banco: str
    cuota_mensual: float
    plazo_inicial: int
    plazo_pendiente: int
    tasa_ea: float
    frech_subsidio: float
    seguro_vida: float
    seguro_incendio: float
    seguro_terremoto: float
    capital_mensual: float
    interes_mensual: float
    saldo_capital: float
    consultor: str
    actividad_economica: str
    abono_efectivo: float = 0
    ingresos: float = 0
    plazos_anos: list[float] | None = None
    tasa_personalizada: float | None = None


def _cell_ref(cell_elem):
    return cell_elem.get('r', '')


def _set_cell_value(cell_elem, value, cell_type='num'):
    """Modifica un elemento <c> del XML: quita formula, pone valor."""
    ns_tag = f'{{{NS}}}'
    # Quitar formula si existe
    f_elem = cell_elem.find(f'{ns_tag}f')
    if f_elem is not None:
        cell_elem.remove(f_elem)
    # Poner/actualizar valor
    v_elem = cell_elem.find(f'{ns_tag}v')
    if v_elem is None:
        v_elem = ET.SubElement(cell_elem, f'{ns_tag}v')
    if cell_type == 'str':
        cell_elem.set('t', 'inlineStr')
        # Remove old v element, use inlineStr
        if v_elem is not None:
            cell_elem.remove(v_elem)
        # Remove existing is element
        is_elem = cell_elem.find(f'{ns_tag}is')
        if is_elem is not None:
            cell_elem.remove(is_elem)
        is_elem = ET.SubElement(cell_elem, f'{ns_tag}is')
        t_elem = ET.SubElement(is_elem, f'{ns_tag}t')
        t_elem.text = str(value)
        # Remove 't' = 'str' if present (inlineStr replaces it)
    else:
        # Numeric
        if 't' in cell_elem.attrib:
            del cell_elem.attrib['t']
        # Remove inlineStr if present
        is_elem = cell_elem.find(f'{ns_tag}is')
        if is_elem is not None:
            cell_elem.remove(is_elem)
        v_elem.text = str(value)


def _find_or_create_cell(sheet_root, cell_ref):
    """Encuentra celda en sheetData o la crea."""
    ns_tag = f'{{{NS}}}'
    col_letter = re.match(r'([A-Z]+)', cell_ref).group(1)
    row_num = int(re.match(r'[A-Z]+(\d+)', cell_ref).group(1))
    
    sheet_data = sheet_root.find(f'{ns_tag}sheetData')
    
    # Find row
    target_row = None
    for row_elem in sheet_data.findall(f'{ns_tag}row'):
        if int(row_elem.get('r', '0')) == row_num:
            target_row = row_elem
            break
    
    if target_row is None:
        target_row = ET.SubElement(sheet_data, f'{ns_tag}row')
        target_row.set('r', str(row_num))
    
    # Find cell
    for cell_elem in target_row.findall(f'{ns_tag}c'):
        if cell_elem.get('r') == cell_ref:
            return cell_elem
    
    # Create cell
    cell_elem = ET.SubElement(target_row, f'{ns_tag}c')
    cell_elem.set('r', cell_ref)
    return cell_elem


def _modify_workbook_xml(
    wb_xml_bytes,
    sheet_index=2,
    new_area="ESTUDIO!$A$1:$R$85",
    active_tab_idx=None,
    window_layout=None,
):
    # FIX 2026-05-12: default new_area era "$A$1:$R$54" (stale). El call site
    # real en crear_estudio pasa "$A$1:$R$85". Si alguien llama sin pasar
    # new_area, antes obtenia rango chico que cortaba el print area en R54.
    """Modifica workbook.xml via REGEX (no XML parser) para preservar namespaces originales.

    Hace:
    - Fija Print_Area de la hoja ESTUDIO (sheet_index)
    - Fija activeTab del workbookView a active_tab_idx (si None, usa sheet_index).
      active_tab_idx es la hoja que queda ACTIVA al abrir (ej. ACTUAL).
    - Fija xWindow/yWindow/windowWidth/windowHeight de workbookView (window_layout
      es dict con esas 4 claves; si None se usa preset mitad derecha 1920x1080).
    - Elimina definedName con referencias rotas ([1]... o #REF!)
    - Elimina <externalReferences>...</externalReferences>
    """
    xml = wb_xml_bytes.decode('utf-8') if isinstance(wb_xml_bytes, bytes) else wb_xml_bytes

    if active_tab_idx is None:
        active_tab_idx = sheet_index

    # Preset: mitad derecha de pantalla 1920x1080 (DPI 96).
    # Excel usa twips (1/1440 in). 1920px @ 96dpi = 20in = 28800 twips.
    if window_layout is None:
        window_layout = {
            'xWindow': 14400,       # mitad pantalla (1920/2 px = ~14400 twips)
            'yWindow': 0,
            'windowWidth': 14400,   # ocupa mitad derecha
            'windowHeight': 15600,  # ~1080px menos barra tareas
        }

    # 1) Actualizar (o insertar) Print_Area de la hoja destino
    def _pa_repl(m):
        attrs = m.group(1)
        return f'<definedName{attrs}>{new_area}</definedName>'

    pa_pattern = re.compile(
        r'<definedName([^>]*name="_xlnm\.Print_Area"[^>]*localSheetId="' + str(sheet_index) + r'"[^>]*)>[^<]*</definedName>'
    )
    if pa_pattern.search(xml):
        xml = pa_pattern.sub(_pa_repl, xml)
    else:
        # La plantilla no trae Print_Area para ESTUDIO -> insertarlo.
        new_dn = (f'<definedName name="_xlnm.Print_Area" '
                  f'localSheetId="{sheet_index}">{new_area}</definedName>')
        if '<definedNames>' in xml:
            xml = xml.replace('<definedNames>', f'<definedNames>{new_dn}', 1)
        elif '<definedNames/>' in xml:
            xml = xml.replace('<definedNames/>', f'<definedNames>{new_dn}</definedNames>', 1)
        else:
            # Insertar un bloque nuevo justo antes de <calcPr o </workbook>
            if '<calcPr' in xml:
                xml = re.sub(r'(<calcPr\b)',
                             f'<definedNames>{new_dn}</definedNames>\\1',
                             xml, count=1)
            else:
                xml = xml.replace('</workbook>',
                                  f'<definedNames>{new_dn}</definedNames></workbook>')

    # 1.b) Fijar activeTab + dimensiones de ventana del workbookView.
    # activeTab apunta a la hoja que quedara activa al abrir (ACTUAL).
    # xWindow/yWindow/windowWidth/windowHeight ubican la ventana en pantalla.
    def _update_workbook_view(tag):
        # Reescribir activeTab
        if re.search(r'\bactiveTab="', tag):
            tag = re.sub(
                r'\bactiveTab="\d+"',
                'activeTab="' + str(active_tab_idx) + '"',
                tag,
                count=1,
            )
        else:
            insert = ' activeTab="' + str(active_tab_idx) + '"'
            if tag.endswith('/>'):
                tag = tag[:-2] + insert + '/>'
            else:
                tag = tag[:-1] + insert + '>'
        # Reescribir/insertar dimensiones de ventana
        for attr in ('xWindow', 'yWindow', 'windowWidth', 'windowHeight'):
            val = str(window_layout.get(attr, 0))
            if re.search(r'\b' + attr + r'="', tag):
                tag = re.sub(r'\b' + attr + r'="[^"]*"', attr + '="' + val + '"', tag, count=1)
            else:
                insert = ' ' + attr + '="' + val + '"'
                if tag.endswith('/>'):
                    tag = tag[:-2] + insert + '/>'
                else:
                    tag = tag[:-1] + insert + '>'
        return tag

    if re.search(r'<workbookView\b', xml):
        xml = re.sub(
            r'<workbookView\b[^>]*?/?>',
            lambda m: _update_workbook_view(m.group(0)),
            xml,
            count=1,
        )
    else:
        # Si no existe workbookView, insertarlo dentro de <bookViews>.
        new_wv = (
            '<workbookView xWindow="' + str(window_layout['xWindow']) + '" '
            'yWindow="' + str(window_layout['yWindow']) + '" '
            'windowWidth="' + str(window_layout['windowWidth']) + '" '
            'windowHeight="' + str(window_layout['windowHeight']) + '" '
            'activeTab="' + str(active_tab_idx) + '"/>'
        )
        if '<bookViews>' in xml:
            xml = xml.replace('<bookViews>', '<bookViews>' + new_wv, 1)
        elif '<bookViews/>' in xml:
            xml = xml.replace('<bookViews/>', '<bookViews>' + new_wv + '</bookViews>', 1)

    # 2) Eliminar definedNames con referencias rotas
    broken_dn = re.compile(
        r'<definedName\b[^>]*>[^<]*(?:\[1\][^<]*|#REF!)[^<]*</definedName>\s*'
    )
    xml = broken_dn.sub('', xml)

    # 3) Eliminar <externalReferences>...</externalReferences>
    xml = re.sub(r'<externalReferences>.*?</externalReferences>\s*', '', xml, flags=re.DOTALL)

    # 4) Eliminar auto-closing <externalReferences/>
    xml = re.sub(r'<externalReferences\s*/>\s*', '', xml)

    # 5) CRITICO: forzar recalculo al abrir Excel.
    # Al cambiar B8 (tasa), B15 (saldo), etc., los valores cacheados de K14,
    # K19, I19 y todas las celdas con formulas quedan obsoletos. Sin forzar
    # recalculo, Excel muestra los cacheados hasta que el usuario pulse F9.
    if 'fullCalcOnLoad' not in xml:
        # Caso A: <calcPr .../> (self-closing) -> agregar atributo
        m = re.search(r'<calcPr([^/>]*)/>', xml)
        if m:
            new_attrs = m.group(1)
            if 'fullCalcOnLoad' not in new_attrs:
                new_attrs = new_attrs.rstrip() + ' fullCalcOnLoad="1"'
            xml = xml.replace(m.group(0), f'<calcPr{new_attrs}/>')
        else:
            # Caso B: <calcPr ...> (abierto) -> agregar atributo antes de >
            m = re.search(r'<calcPr([^>]*)>', xml)
            if m:
                new_attrs = m.group(1)
                if 'fullCalcOnLoad' not in new_attrs:
                    new_attrs = new_attrs.rstrip() + ' fullCalcOnLoad="1"'
                xml = xml.replace(m.group(0), f'<calcPr{new_attrs}>')
            else:
                # Caso C: no hay calcPr -> insertar antes de </workbook>
                xml = xml.replace('</workbook>', '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')

    return xml.encode('utf-8')


def _remove_external_links(wb_xml_bytes, rels_xml_bytes):
    """Quita referencias a external links y connections del workbook."""
    # Remove from workbook.xml
    wb_root = ET.fromstring(wb_xml_bytes)
    ns_tag = f'{{{NS}}}'
    
    # Remove externalReferences element
    for elem in wb_root.findall(f'{ns_tag}externalReferences'):
        wb_root.remove(elem)
    
    wb_out = ET.tostring(wb_root, xml_declaration=True, encoding='UTF-8')
    
    # Remove from rels
    ns_rels = 'http://schemas.openxmlformats.org/package/2006/relationships'
    rels_root = ET.fromstring(rels_xml_bytes)
    to_remove = []
    for rel in rels_root:
        rtype = rel.get('Type', '')
        if 'externalLink' in rtype or 'connections' in rtype:
            to_remove.append(rel)
    for r in to_remove:
        rels_root.remove(r)
    
    rels_out = ET.tostring(rels_root, xml_declaration=True, encoding='UTF-8')
    return wb_out, rels_out


def _fix_sheet_view_xml(sheet_xml_bytes, is_estudio: bool, is_actual: bool = False):
    """Ajusta <sheetView> de un worksheet XML sin alterar su contenido.

    - En ACTUAL (Jose 2026-04-18): fija tabSelected='1', zoomScale='145',
      zoomScaleNormal='145'. Sin view=pageBreakPreview (vista normal).
      ACTUAL es la hoja activa al abrir el archivo.
    - En ESTUDIO: fija view='pageBreakPreview', zoomScale='85',
      zoomScalePageLayoutView='70'. SIN tabSelected (ya no es la activa).
      Preserva layout de impresion para cuando el usuario imprima.
    - En cualquier otra hoja: elimina tabSelected para evitar "selecciones
      multiples" al copiar (hojas agrupadas).

    Trabaja con regex sobre el XML crudo para NO pasar por openpyxl y asi
    preservar drawings/imagenes/logo embebidos en el template.
    """
    xml = sheet_xml_bytes.decode('utf-8') if isinstance(sheet_xml_bytes, bytes) else sheet_xml_bytes

    def _update_view_attrs(m):
        tag = m.group(0)
        # Eliminar atributos que vamos a reescribir
        tag = re.sub(r'\s+view="[^"]*"', '', tag)
        tag = re.sub(r'\s+tabSelected="[^"]*"', '', tag)
        tag = re.sub(r'\s+zoomScale="[^"]*"', '', tag)
        tag = re.sub(r'\s+zoomScaleNormal="[^"]*"', '', tag)
        tag = re.sub(r'\s+zoomScalePageLayoutView="[^"]*"', '', tag)
        if is_actual:
            extra = ' tabSelected="1" zoomScale="145" zoomScaleNormal="145"'
        elif is_estudio:
            extra = ' view="pageBreakPreview" zoomScale="85" zoomScalePageLayoutView="70"'
        else:
            extra = ''
        if not extra:
            return tag
        # Insertar antes de '>' o '/>'
        if tag.endswith('/>'):
            return tag[:-2] + extra + '/>'
        return tag[:-1] + extra + '>'

    xml = re.sub(r'<sheetView\b[^>]*?/?>', _update_view_attrs, xml, count=1)
    return xml.encode('utf-8')


def _detectar_estudio(wb_xml_bytes, rels_xml_bytes):
    """Detecta posicion (0-based) y filename fisico de la hoja ESTUDIO.

    Lee workbook.xml y workbook.xml.rels para resolver:
    - estudio_idx: posicion en <sheets> (para activeTab y localSheetId)
    - estudio_sheet_file: 'xl/worksheets/sheetN.xml' real del ESTUDIO

    Evita hardcodes que se rompen si el template reordena sus hojas
    (BUG B, Jose 2026-04-17).
    """
    wb = wb_xml_bytes.decode('utf-8') if isinstance(wb_xml_bytes, bytes) else wb_xml_bytes
    rels = rels_xml_bytes.decode('utf-8') if isinstance(rels_xml_bytes, bytes) else rels_xml_bytes

    # 1) Encontrar la hoja ESTUDIO y la hoja BD en el listado <sheets>
    #    Regex permisivo: matchea <sheet .../> y <sheet ...> (ambas formas).
    estudio_idx = None
    estudio_rid = None
    bd_idx = None
    nombres_vistos = []
    for idx, m in enumerate(re.finditer(r'<sheet\b[^>]*?>', wb)):
        tag = m.group(0)
        name_m = re.search(r'\bname="([^"]*)"', tag)
        name_val = name_m.group(1).strip() if name_m else ''
        nombres_vistos.append(name_val)
        name_upper = name_val.upper()
        if name_upper == 'ESTUDIO' and estudio_idx is None:
            estudio_idx = idx
            # r:id puede venir con o sin prefijo de ns segun como lo escriba Excel
            rid_m = re.search(r'(?:r:id|r:Id|relationshipId)="([^"]*)"', tag)
            estudio_rid = rid_m.group(1) if rid_m else None
        elif name_upper == 'BD' and bd_idx is None:
            bd_idx = idx

    if estudio_idx is None or not estudio_rid:
        raise ValueError(
            "No se encontro hoja ESTUDIO en workbook.xml. "
            f"Hojas detectadas: {nombres_vistos}"
        )

    # BUG B (Jose 2026-04-17): compensar el reordenamiento que hace
    # ocultar_hoja_bd() POSTERIOR a este paso. Si BD es la PRIMERA hoja del
    # template, sera movida al final -> el idx de ESTUDIO baja en 1.
    if bd_idx is not None and bd_idx == 0 and bd_idx < estudio_idx:
        estudio_idx -= 1

    # 2) Resolver rId -> Target en workbook.xml.rels
    target = None
    for m in re.finditer(r'<Relationship\b[^>]*?>', rels):
        a = m.group(0)
        id_m = re.search(r'\bId="([^"]*)"', a)
        tgt_m = re.search(r'\bTarget="([^"]*)"', a)
        if id_m and id_m.group(1) == estudio_rid and tgt_m:
            target = tgt_m.group(1)
            break

    if not target:
        raise ValueError(f"No se encontro Target para rId={estudio_rid}")

    # Normalizar a ruta completa dentro del zip: 'xl/worksheets/sheetN.xml'
    # Manejar '/xl/...', 'xl/...' y 'worksheets/...' robustamente.
    target = target.lstrip('/')
    if not target.startswith('xl/'):
        target = 'xl/' + target

    return estudio_idx, target


def _detectar_actual(wb_xml_bytes, rels_xml_bytes):
    """Detecta posicion (0-based visible) y filename fisico de la hoja ACTUAL.

    Analogo a _detectar_estudio pero para la hoja ACTUAL (dashboard del
    cliente donde se analiza comparativamente con el extracto).
    """
    wb = wb_xml_bytes.decode('utf-8') if isinstance(wb_xml_bytes, bytes) else wb_xml_bytes
    rels = rels_xml_bytes.decode('utf-8') if isinstance(rels_xml_bytes, bytes) else rels_xml_bytes

    actual_idx = None
    actual_rid = None
    bd_idx = None
    nombres_vistos = []
    for idx, m in enumerate(re.finditer(r'<sheet\b[^>]*?>', wb)):
        tag = m.group(0)
        name_m = re.search(r'\bname="([^"]*)"', tag)
        name_val = name_m.group(1).strip() if name_m else ''
        nombres_vistos.append(name_val)
        name_upper = name_val.upper()
        if name_upper == 'ACTUAL' and actual_idx is None:
            actual_idx = idx
            rid_m = re.search(r'(?:r:id|r:Id|relationshipId)="([^"]*)"', tag)
            actual_rid = rid_m.group(1) if rid_m else None
        elif name_upper == 'BD' and bd_idx is None:
            bd_idx = idx

    if actual_idx is None or not actual_rid:
        raise ValueError(
            "No se encontro hoja ACTUAL en workbook.xml. "
            f"Hojas detectadas: {nombres_vistos}"
        )

    # Misma compensacion por reordenamiento de BD (ocultamiento la mueve al final)
    if bd_idx is not None and bd_idx == 0 and bd_idx < actual_idx:
        actual_idx -= 1

    target = None
    for m in re.finditer(r'<Relationship\b[^>]*?>', rels):
        a = m.group(0)
        id_m = re.search(r'\bId="([^"]*)"', a)
        tgt_m = re.search(r'\bTarget="([^"]*)"', a)
        if id_m and id_m.group(1) == actual_rid and tgt_m:
            target = tgt_m.group(1)
            break

    if not target:
        raise ValueError(f"No se encontro Target para rId={actual_rid}")

    target = target.lstrip('/')
    if not target.startswith('xl/'):
        target = 'xl/' + target

    return actual_idx, target


class ExcelPopulator:
    """Crea copias standalone de PESOS.xlsx manipulando ZIP directamente."""

    def __init__(self, template_path: str):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template no encontrado: {template_path}")
        self.template_path = template_path

    def crear_estudio(self, datos: DatosClienteExcel, carpeta_salida: str, fecha: str | None = None) -> str:
        if fecha is None:
            fecha = datetime.now().strftime("%d.%m.%y")

        nombre_upper = datos.nombre.upper().strip()
        # Naming canonico (Jose 2026-04-17): guion medio entre nombre y fecha
        filename = f"ESTUDIO {nombre_upper}-{fecha}.xlsx"
        output_path = os.path.join(carpeta_salida, filename)
        os.makedirs(carpeta_salida, exist_ok=True)

        # Build values dict from datos
        values = {}
        for cell_ref, (ctype, field) in INDEX_CELLS.items():
            values[cell_ref] = (ctype, getattr(datos, field))
        for cell_ref, (ctype, field) in DIRECT_CELLS.items():
            val = getattr(datos, field)
            if field == 'tasa_ea' and datos.tasa_personalizada is not None:
                val = datos.tasa_personalizada
            values[cell_ref] = (ctype, val)
        
        if datos.plazos_anos:
            for i, plazo in enumerate(datos.plazos_anos[:6]):
                values[PLAZO_CELLS[i]] = ('num', plazo)

        # Manipulate ZIP directly
        ns_tag = f'{{{NS}}}'
        
        with zipfile.ZipFile(self.template_path, 'r') as zin:
            # Pre-scan: detectar posicion y filename fisico de ESTUDIO
            # (evita hardcodes sheet_index=2 y sheet3.xml que se rompen
            # si el template reordena sus hojas — BUG B, Jose 2026-04-17)
            _wb_pre = zin.read('xl/workbook.xml')
            _rels_pre = zin.read('xl/_rels/workbook.xml.rels')
            estudio_idx, estudio_sheet_file = _detectar_estudio(_wb_pre, _rels_pre)
            actual_idx, actual_sheet_file = _detectar_actual(_wb_pre, _rels_pre)
            # Validar que el filename calculado existe en el zip
            if estudio_sheet_file not in zin.namelist():
                _zip_sheets = [n for n in zin.namelist()
                               if n.startswith('xl/worksheets/sheet')]
                raise ValueError(
                    f"estudio_sheet_file='{estudio_sheet_file}' no existe en el zip. "
                    f"Sheets disponibles: {_zip_sheets}"
                )
            if actual_sheet_file not in zin.namelist():
                _zip_sheets = [n for n in zin.namelist()
                               if n.startswith('xl/worksheets/sheet')]
                raise ValueError(
                    f"actual_sheet_file='{actual_sheet_file}' no existe en el zip. "
                    f"Sheets disponibles: {_zip_sheets}"
                )

            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    
                    if item.filename == actual_sheet_file:
                        # ACTUAL sheet: replace INDEX/MATCH with direct values.
                        # FIX 2026-05-12: usa actual_sheet_file detectado dinamicamente
                        # (era hardcoded 'xl/worksheets/sheet2.xml' — si template reordena
                        # hojas, el detector arriba lo capturaba pero este bloque NO
                        # ejecutaba, dejando INDEX/MATCH sin reemplazar).
                        root = ET.fromstring(data)
                        sheet_data = root.find(f'{ns_tag}sheetData')

                        for row_elem in sheet_data.findall(f'{ns_tag}row'):
                            for cell_elem in row_elem.findall(f'{ns_tag}c'):
                                ref = cell_elem.get('r', '')
                                if ref in values:
                                    ctype, val = values[ref]
                                    _set_cell_value(cell_elem, val, ctype)

                        data = ET.tostring(root, xml_declaration=True, encoding='UTF-8')
                        data = _fix_xml_namespaces(data)
                    
                    elif item.filename == 'xl/workbook.xml':
                        # Fix print area + limpiar definedNames rotos + remove externalReferences
                        # (usa regex, no XML parser, para preservar namespaces originales)
                        # activeTab apunta a ACTUAL (hoja activa al abrir).
                        # sheet_index sigue apuntando a ESTUDIO (para Print_Area).
                        data = _modify_workbook_xml(
                            data,
                            sheet_index=estudio_idx,
                            new_area="ESTUDIO!$A$1:$R$85",
                            active_tab_idx=actual_idx,
                        )
                        # Asegurar que prefijos usados (ej. xcalcf) esten declarados en root
                        data = _fix_xml_namespaces(data)

                    elif item.filename == 'xl/_rels/workbook.xml.rels':
                        # Remove external link, connection and calcChain rels
                        ns_rels = 'http://schemas.openxmlformats.org/package/2006/relationships'
                        ET.register_namespace('', ns_rels)
                        rels_root = ET.fromstring(data)
                        to_remove = [r for r in rels_root
                                     if 'externalLink' in r.get('Type','')
                                     or 'connections' in r.get('Type','')
                                     or 'calcChain' in r.get('Type','')
                                     or 'calcChain' in r.get('Target','')]
                        for r in to_remove:
                            rels_root.remove(r)
                        data = ET.tostring(rels_root, xml_declaration=True, encoding='UTF-8')
                        data = _fix_xml_namespaces(data)

                    elif 'externalLinks' in item.filename:
                        continue  # Skip external link files entirely
                    
                    elif item.filename == 'xl/connections.xml':
                        continue  # Skip connections file
                    
                    elif item.filename == 'xl/queryTables/queryTable1.xml':
                        continue  # Skip query tables
                    
                    elif item.filename == '[Content_Types].xml':
                        # Remove content type entries for deleted files
                        ct_root = ET.fromstring(data)
                        ct_ns = 'http://schemas.openxmlformats.org/package/2006/content-types'
                        ET.register_namespace('', ct_ns)
                        to_remove = [e for e in ct_root
                                     if 'externalLink' in e.get('PartName','')
                                     or 'connections' in e.get('PartName','')
                                     or 'queryTable' in e.get('PartName','')
                                     or 'calcChain' in e.get('PartName','')]
                        for e in to_remove:
                            ct_root.remove(e)
                        data = ET.tostring(ct_root, xml_declaration=True, encoding='UTF-8')
                        data = _fix_xml_namespaces(data)

                    elif item.filename == 'xl/calcChain.xml':
                        continue  # Skip stale calcChain, Excel will rebuild it

                    # FIX CRITICO: borrar valores cacheados de CUALQUIER hoja con
                    # formulas. Los valores <v>...</v> dentro de celdas que tienen
                    # <f>...</f> son cache del cliente ANTERIOR del template.
                    # Al borrarlos, Excel NO TIENE opcion: recalcula todo al abrir.
                    # Esto complementa fullCalcOnLoad="1" y garantiza recalculo
                    # de las 6 opciones (hojas PESOS 1-6) y ESTUDIO.
                    if item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
                        data_str = data.decode('utf-8') if isinstance(data, bytes) else data
                        # Remover <v>...</v> que esta DENTRO de una celda que tambien tiene <f>
                        # Pattern: <c ...><f>...</f>...<v>VALOR</v></c>
                        data_str = re.sub(
                            r'(<c\b[^>]*>\s*(?:<f\b[^/>]*(?:/>|>[^<]*</f>))\s*)<v>[^<]*</v>',
                            r'\1',
                            data_str,
                        )
                        data = data_str.encode('utf-8')

                        # Fix sheetView (Jose 2026-04-17): ESTUDIO con
                        # pageBreakPreview + tabSelected=1; otras hojas sin
                        # tabSelected (evita error "selecciones multiples" al
                        # copiar por hojas agrupadas). Trabaja en XML crudo
                        # para preservar drawings/logo del template.
                        is_estudio = (item.filename == estudio_sheet_file)
                        is_actual = (item.filename == actual_sheet_file)
                        data = _fix_sheet_view_xml(
                            data, is_estudio=is_estudio, is_actual=is_actual
                        )

                    zout.writestr(item, data)

        return output_path

    def crear_desde_bd(self, credito_id: str, carpeta_salida: str, fecha: str | None = None) -> str:
        """Busca cliente en BD y crea estudio."""
        import openpyxl
        wb = openpyxl.load_workbook(self.template_path, data_only=True)
        ws = wb['BD']
        
        fila = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if val is not None and str(val).strip() == str(credito_id).strip():
                fila = row
                break
        
        if fila is None:
            wb.close()
            raise ValueError(f"Credito '{credito_id}' no encontrado en BD")

        # Validacion: solo procesamos creditos en PESOS (no UVR)
        amortizacion = ws.cell(row=fila, column=19).value
        if amortizacion:
            amort_upper = str(amortizacion).upper().strip()
            if 'UVR' in amort_upper:
                wb.close()
                raise ValueError(
                    f"Credito '{credito_id}' es en UVR (amortizacion: {amortizacion}). "
                    "Actualmente solo se procesan creditos en PESOS."
                )

        def get(col):
            v = ws.cell(row=fila, column=col).value
            return v if v is not None else 0
        
        tasa_raw = to_float(get(24))
        tasa_ea = tasa_raw / 100 if tasa_raw > 1 else tasa_raw
        
        plazo_pendiente = to_int(get(23))

        # PLAZOS DINAMICOS (confirmado por Jose): SIEMPRE inferiores al plazo
        # pendiente del cliente. Si plazo_pend >= 14 anos, usa defaults
        # [13.5, 12, 11, 10, 9, 8.5]. Si es menor, escalonado decreciente.
        # El analista puede pasar plazos_manuales para overridear.
        plazos_anos_default = _calcular_plazos_dinamicos(plazo_pendiente)

        datos = DatosClienteExcel(
            credito_id=credito_id,
            nombre=str(get(1)),
            banco=str(get(6)),
            cuota_mensual=to_float(get(21)),
            plazo_inicial=to_int(get(22)),
            plazo_pendiente=plazo_pendiente,
            tasa_ea=tasa_ea,
            frech_subsidio=to_float(get(25)),
            seguro_vida=to_float(get(26)),
            seguro_incendio=to_float(get(27)),
            seguro_terremoto=to_float(get(28)),
            capital_mensual=to_float(get(29)),
            interes_mensual=to_float(get(30)),
            saldo_capital=to_float(get(31)),
            consultor=str(get(13)),
            actividad_economica=str(get(34)),
            abono_efectivo=to_float(get(32)),
            ingresos=to_float(get(33)),
            plazos_anos=plazos_anos_default,
        )
        wb.close()
        return self.crear_estudio(datos, carpeta_salida, fecha)


class PDFExporter:
    """Exporta ESTUDIO sheet a PDF via LibreOffice headless."""

    @staticmethod
    def _prepare_temp_xlsx(excel_path, temp_xlsx):
        """Prepara copia temporal via ZIP/XML: oculta hojas + limpia filas 55+ de ESTUDIO.

        Aplica _fix_xml_namespaces a los XML modificados para que Excel COM
        no los rechace.
        """
        ns = f'{{{NS}}}'
        with zipfile.ZipFile(excel_path, 'r') as zin:
            with zipfile.ZipFile(temp_xlsx, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == 'xl/workbook.xml':
                        root = ET.fromstring(data)
                        for sheet in root.findall(f'.//{ns}sheet'):
                            name = sheet.get('name')
                            if name == 'ESTUDIO':
                                if 'state' in sheet.attrib:
                                    del sheet.attrib['state']
                            else:
                                sheet.set('state', 'hidden')
                        data = ET.tostring(root, xml_declaration=True, encoding='UTF-8')
                        data = _fix_xml_namespaces(data)
                    elif item.filename == 'xl/worksheets/sheet3.xml':
                        # TODO 2026-05-12: hardcoded sheet3.xml para hoja ESTUDIO.
                        # Si template reordena hojas, esto deja de funcionar.
                        # Cuando se active la feature PDFExporter, refactor para
                        # usar _detectar_estudio() como hace ExcelPopulator.
                        # Por ahora no es ruta productiva (PDFExporter no se usa
                        # en pipeline_davivienda — solo manual via
                        # generar_estudio_completo con generar_pdf=True).
                        root = ET.fromstring(data)
                        sheet_data = root.find(f'{ns}sheetData')
                        rows_to_remove = []
                        for row_elem in sheet_data.findall(f'{ns}row'):
                            row_num = int(row_elem.get('r', '0'))
                            if row_num >= 55:
                                rows_to_remove.append(row_elem)
                        for row_elem in rows_to_remove:
                            sheet_data.remove(row_elem)
                        merge_cells = root.find(f'{ns}mergeCells')
                        if merge_cells is not None:
                            merges_to_remove = []
                            for mc in merge_cells.findall(f'{ns}mergeCell'):
                                ref = mc.get('ref', '')
                                parts = ref.split(':')
                                for part in parts:
                                    row_match = re.search(r'(\d+)', part)
                                    if row_match and int(row_match.group(1)) >= 55:
                                        merges_to_remove.append(mc)
                                        break
                            for mc in merges_to_remove:
                                merge_cells.remove(mc)
                            remaining = merge_cells.findall(f'{ns}mergeCell')
                            if remaining:
                                merge_cells.set('count', str(len(remaining)))
                            else:
                                root.remove(merge_cells)
                        data = ET.tostring(root, xml_declaration=True, encoding='UTF-8')
                        data = _fix_xml_namespaces(data)
                    zout.writestr(item, data)

    @staticmethod
    def _trim_pdf(pdf_path, max_pages=2):
        """Recorta PDF a max_pages usando PyMuPDF."""
        import fitz
        doc = fitz.open(pdf_path)
        if doc.page_count > max_pages:
            pages_to_delete = list(range(max_pages, doc.page_count))
            doc.delete_pages(pages_to_delete)
            temp_path = pdf_path + ".tmp"
            doc.save(temp_path)
            doc.close()
            os.replace(temp_path, pdf_path)
        else:
            doc.close()

    @staticmethod
    def _exportar_con_excel(temp_xlsx, pdf_final):
        """Exporta a PDF usando Excel de Office 365 via COM (win32com).

        Requiere pywin32 y Office 365/2019+ instalado en el PC.
        Si falla, lanza RuntimeError con diagnostico claro.
        """
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            raise RuntimeError(
                "pywin32 no esta instalado. Instalalo con: py -m pip install pywin32"
            )

        excel_app = None
        wb = None
        try:
            # Inicializar COM para este hilo
            pythoncom.CoInitialize()

            try:
                excel_app = win32com.client.DispatchEx("Excel.Application")
            except Exception as e:
                raise RuntimeError(
                    f"No se pudo iniciar Excel. Verifica que Office 365 este instalado. "
                    f"Detalle: {e}"
                )

            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = False

            abs_xlsx = os.path.abspath(temp_xlsx)
            if not os.path.exists(abs_xlsx):
                raise RuntimeError(f"El archivo temporal no existe: {abs_xlsx}")

            try:
                wb = excel_app.Workbooks.Open(
                    abs_xlsx,
                    UpdateLinks=0,      # No actualizar links externos
                    ReadOnly=True,      # Abrir solo-lectura (mas rapido)
                    IgnoreReadOnlyRecommended=True,
                    CorruptLoad=0,      # Modo normal (no recuperacion)
                )
            except Exception as e:
                raise RuntimeError(
                    f"Excel no pudo abrir el archivo temporal.\n"
                    f"  Archivo: {abs_xlsx}\n"
                    f"  Detalle COM: {e}\n"
                    f"  Posibles causas: archivo corrupto, Excel protegido, permisos"
                )

            try:
                ws = wb.Sheets("ESTUDIO")
            except Exception as e:
                raise RuntimeError(f"No se encontro la hoja ESTUDIO: {e}")

            # Forzar recalculo antes de exportar
            try:
                excel_app.CalculateFull()
            except Exception:
                pass

            # Area de impresion
            ws.PageSetup.PrintArea = "$A$1:$R$54"

            # Exportar
            try:
                ws.ExportAsFixedFormat(
                    Type=0,  # xlTypePDF
                    Filename=os.path.abspath(pdf_final),
                    Quality=0,  # xlQualityStandard
                    IncludeDocProperties=False,
                    IgnorePrintAreas=False,
                    OpenAfterPublish=False,
                )
            except Exception as e:
                raise RuntimeError(f"ExportAsFixedFormat fallo: {e}")

        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel_app is not None:
                    excel_app.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    @staticmethod
    def _exportar_con_libreoffice(temp_xlsx, output_dir):
        """Exporta a PDF usando LibreOffice headless (fallback)."""
        cmd = ["libreoffice", "--headless", "--calc", "--convert-to", "pdf", "--outdir", output_dir, temp_xlsx]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice fallo: {result.stderr}")
        return os.path.join(output_dir, os.path.splitext(os.path.basename(temp_xlsx))[0] + ".pdf")

    @staticmethod
    def exportar(excel_path, output_dir=None, metodo="excel"):
        """Exporta solo ESTUDIO a PDF.
        metodo: 'excel' (Office 365 via COM) o 'libreoffice' (headless).
        """
        if output_dir is None:
            output_dir = os.path.dirname(excel_path)
        excel_path = os.path.abspath(excel_path)
        output_dir = os.path.abspath(output_dir)
        os.makedirs(output_dir, exist_ok=True)

        pdf_name = os.path.splitext(os.path.basename(excel_path))[0] + ".pdf"
        pdf_final = os.path.join(output_dir, pdf_name)

        temp_dir = tempfile.mkdtemp()
        temp_xlsx = os.path.join(temp_dir, os.path.basename(excel_path))
        try:
            PDFExporter._prepare_temp_xlsx(excel_path, temp_xlsx)
            if metodo == "excel":
                PDFExporter._exportar_con_excel(temp_xlsx, pdf_final)
            elif metodo == "libreoffice":
                pdf_gen = PDFExporter._exportar_con_libreoffice(temp_xlsx, output_dir)
                if pdf_gen != pdf_final and os.path.exists(pdf_gen):
                    if os.path.exists(pdf_final):
                        os.remove(pdf_final)
                    os.rename(pdf_gen, pdf_final)
            else:
                raise ValueError("Metodo desconocido: " + str(metodo))
            if not os.path.exists(pdf_final):
                raise FileNotFoundError("PDF no generado: " + pdf_final)
            PDFExporter._trim_pdf(pdf_final, max_pages=2)
            return pdf_final
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)


def generar_estudio_completo(datos, template_path, carpeta_salida, generar_pdf=True, fecha=None):
    populator = ExcelPopulator(template_path)
    excel_path = populator.crear_estudio(datos, carpeta_salida, fecha)
    result = {"excel": excel_path, "pdf": None}
    if generar_pdf:
        try:
            pdf_path = PDFExporter.exportar(excel_path, carpeta_salida)
            result["pdf"] = pdf_path
        except Exception as e:
            print("Error generando PDF: " + str(e))
    return result
